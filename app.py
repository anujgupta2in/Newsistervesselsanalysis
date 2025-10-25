import streamlit as st
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
from datetime import datetime
import re
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from io import BytesIO

st.set_page_config(page_title="Vessel Maintenance Job Analysis", layout="wide")

st.title("Vessel Maintenance Job Analysis")
st.markdown("Upload CSV or Excel files to analyze vessel maintenance jobs")

def extract_date_from_filename(filename):
    match = re.search(r'(\d{8})', filename)
    if match:
        date_str = match.group(1)
        try:
            date_obj = datetime.strptime(date_str, '%d%m%Y')
            return date_obj.strftime('%d-%m-%Y')
        except:
            return date_str
    return "Unknown"

def clean_dataframe(df):
    for col in df.columns:
        if df[col].dtype == 'object':
            df[col] = df[col].astype(str).str.strip()
    return df

def process_file(file, filename):
    if filename.endswith('.csv'):
        df = pd.read_csv(file, encoding='utf-8-sig')
    else:
        df = pd.read_excel(file)
    
    df = clean_dataframe(df)
    
    if len(df.columns) > 1:
        df['Is_Critical'] = df.iloc[:, 1] == 'C'
    else:
        df['Is_Critical'] = False
    
    return df

uploaded_files = st.file_uploader(
    "Upload CSV or Excel files", 
    type=['csv', 'xlsx', 'xls'], 
    accept_multiple_files=True
)

if uploaded_files:
    all_data = []
    file_summaries = []
    
    for uploaded_file in uploaded_files:
        filename = uploaded_file.name
        df = process_file(uploaded_file, filename)
        
        file_date = extract_date_from_filename(filename)
        
        vessel_col = 'Vessel' if 'Vessel' in df.columns else None
        machinery_col = 'Machinery Location' if 'Machinery Location' in df.columns else None
        
        total_jobs = len(df)
        critical_jobs = df['Is_Critical'].sum()
        
        file_info = {
            'filename': filename,
            'date': file_date,
            'total_jobs': total_jobs,
            'critical_jobs': critical_jobs,
            'vessel_col': vessel_col,
            'machinery_col': machinery_col
        }
        file_summaries.append(file_info)
        
        df['Source_File'] = filename
        df['File_Date'] = file_date
        all_data.append(df)
    
    combined_df = pd.concat(all_data, ignore_index=True)
    
    st.header("📊 File Summary")
    
    summary_data = []
    for fs in file_summaries:
        summary_data.append({
            'File Name': fs['filename'],
            'Date': fs['date'],
            'Total Jobs': fs['total_jobs'],
            'Critical Jobs': fs['critical_jobs'],
            'Non-Critical Jobs': fs['total_jobs'] - fs['critical_jobs']
        })
    
    summary_df = pd.DataFrame(summary_data)
    st.dataframe(summary_df, use_container_width=True)
    
    st.header("🚢 File-wise Analysis")
    
    if 'Vessel' in combined_df.columns and 'Source_File' in combined_df.columns:
        file_summary = combined_df.groupby(['Source_File', 'Vessel', 'File_Date']).agg(
            Total_Jobs=('Is_Critical', 'count'),
            Critical_Jobs=('Is_Critical', 'sum')
        ).reset_index()
        
        file_summary['Non_Critical_Jobs'] = file_summary['Total_Jobs'] - file_summary['Critical_Jobs']
        
        st.dataframe(file_summary, use_container_width=True)
        
        fig = px.bar(file_summary, x='Source_File', y=['Critical_Jobs', 'Non_Critical_Jobs'],
                     title='Job Distribution by File',
                     labels={'value': 'Number of Jobs', 'variable': 'Job Type'},
                     barmode='stack',
                     color_discrete_map={'Critical_Jobs': '#ff6b6b', 'Non_Critical_Jobs': '#4ecdc4'})
        st.plotly_chart(fig, use_container_width=True)
    else:
        st.warning("No 'Vessel' or 'Source_File' column found in the uploaded files")
    
    st.header("⚖️ File Comparison by Machinery")
    
    if 'Vessel' in combined_df.columns and 'Machinery Location' in combined_df.columns and 'Source_File' in combined_df.columns:
        files = combined_df['Source_File'].unique()
        
        if len(files) >= 2:
            file_machinery = combined_df.groupby(['Source_File', 'Machinery Location']).size().reset_index(name='Job_Count')
            
            pivot_table = file_machinery.pivot(index='Machinery Location', columns='Source_File', values='Job_Count').fillna(0)
            
            file_cols = list(pivot_table.columns)
            
            for i in range(len(file_cols)):
                for j in range(i + 1, len(file_cols)):
                    diff_col_name = f'Difference ({file_cols[i]} - {file_cols[j]})'
                    pivot_table[diff_col_name] = pivot_table[file_cols[i]] - pivot_table[file_cols[j]]
            
            pivot_table = pivot_table.sort_index()
            pivot_table_display = pivot_table.reset_index()
            
            missing_items = []
            diff_items = []
            
            for idx, row in pivot_table.iterrows():
                has_missing = False
                has_diff = False
                
                for i, file_col in enumerate(file_cols):
                    other_files = [f for f in file_cols if f != file_col]
                    for other_file in other_files:
                        if (row[file_col] > 0 and row[other_file] == 0) or (row[file_col] == 0 and row[other_file] > 0):
                            has_missing = True
                        elif row[file_col] > 0 and row[other_file] > 0 and row[file_col] != row[other_file]:
                            has_diff = True
                
                if has_missing:
                    missing_items.append(idx)
                if has_diff:
                    diff_items.append(idx)
            
            def highlight_comparison(row):
                styles = [''] * len(row)
                
                is_missing = False
                is_diff = False
                
                if 'Machinery Location' in row.index:
                    machinery_name = row['Machinery Location']
                    
                    for i, file_col in enumerate(file_cols):
                        file_val = row[file_col]
                        other_files = [f for f in file_cols if f != file_col]
                        for other_file in other_files:
                            other_val = row[other_file]
                            
                            if (file_val > 0 and other_val == 0) or (file_val == 0 and other_val > 0):
                                is_missing = True
                            elif file_val > 0 and other_val > 0 and file_val != other_val:
                                is_diff = True
                    
                    if is_missing:
                        machinery_idx = row.index.get_loc('Machinery Location')
                        styles[machinery_idx] = 'background-color: #FFB6C1'
                        
                        for file_col in file_cols:
                            file_idx = row.index.get_loc(file_col)
                            styles[file_idx] = 'background-color: #FFB6C1'
                    
                    elif is_diff:
                        machinery_idx = row.index.get_loc('Machinery Location')
                        styles[machinery_idx] = 'background-color: #FFFF99'
                
                for col_name in row.index:
                    if 'Difference' in col_name:
                        diff_idx = row.index.get_loc(col_name)
                        if row[col_name] != 0:
                            styles[diff_idx] = 'background-color: #FFFF99'
                
                return styles
            
            def format_number(val):
                if isinstance(val, (int, float)):
                    return f"{val:.0f}"
                return val
            
            styled_table = pivot_table_display.style.apply(highlight_comparison, axis=1).format(format_number)
            
            st.dataframe(styled_table, use_container_width=True)
            
            def create_excel_with_formatting():
                output = BytesIO()
                with pd.ExcelWriter(output, engine='openpyxl') as writer:
                    pivot_table_display.to_excel(writer, sheet_name='Vessel Comparison', index=False)
                    workbook = writer.book
                    worksheet = writer.sheets['Vessel Comparison']
                    
                    pink_fill = PatternFill(start_color='FFB6C1', end_color='FFB6C1', fill_type='solid')
                    yellow_fill = PatternFill(start_color='FFFF99', end_color='FFFF99', fill_type='solid')
                    
                    for row_idx, row in enumerate(pivot_table_display.itertuples(index=False), start=2):
                        is_missing = False
                        is_diff = False
                        
                        machinery_name = row[0]
                        
                        for i, file_col in enumerate(file_cols):
                            file_val = pivot_table_display.loc[row_idx - 2, file_col]
                            other_files = [f for f in file_cols if f != file_col]
                            for other_file in other_files:
                                other_val = pivot_table_display.loc[row_idx - 2, other_file]
                                
                                if (file_val > 0 and other_val == 0) or (file_val == 0 and other_val > 0):
                                    is_missing = True
                                elif file_val > 0 and other_val > 0 and file_val != other_val:
                                    is_diff = True
                        
                        if is_missing:
                            for col_idx in range(1, len(pivot_table_display.columns) + 1):
                                cell = worksheet.cell(row=row_idx, column=col_idx)
                                cell.fill = pink_fill
                        elif is_diff:
                            machinery_cell = worksheet.cell(row=row_idx, column=1)
                            machinery_cell.fill = yellow_fill
                            
                            for col_name in pivot_table_display.columns:
                                if 'Difference' in col_name:
                                    col_idx = pivot_table_display.columns.get_loc(col_name) + 1
                                    diff_val = pivot_table_display.loc[row_idx - 2, col_name]
                                    if diff_val != 0:
                                        cell = worksheet.cell(row=row_idx, column=col_idx)
                                        cell.fill = yellow_fill
                    
                    for column in worksheet.columns:
                        max_length = 0
                        column_letter = column[0].column_letter
                        for cell in column:
                            try:
                                if len(str(cell.value)) > max_length:
                                    max_length = len(str(cell.value))
                            except:
                                pass
                        adjusted_width = min(max_length + 2, 50)
                        worksheet.column_dimensions[column_letter].width = adjusted_width
                
                output.seek(0)
                return output.getvalue()
            
            excel_data = create_excel_with_formatting()
            st.download_button(
                label="📥 Download File Comparison as Excel",
                data=excel_data,
                file_name="File_Comparison_Analysis.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
            
            col1, col2 = st.columns(2)
            
            with col1:
                with st.expander("🔍 View Missing Items", expanded=False):
                    if missing_items:
                        missing_df = pivot_table.loc[missing_items].reset_index()
                        
                        def highlight_missing(row):
                            styles = [''] * len(row)
                            
                            if 'Machinery Location' in row.index:
                                machinery_idx = row.index.get_loc('Machinery Location')
                                styles[machinery_idx] = 'background-color: #FFB6C1'
                                
                                for file_col in file_cols:
                                    file_idx = row.index.get_loc(file_col)
                                    styles[file_idx] = 'background-color: #FFB6C1'
                                
                                for col_name in row.index:
                                    if 'Difference' in col_name:
                                        diff_idx = row.index.get_loc(col_name)
                                        if row[col_name] != 0:
                                            styles[diff_idx] = 'background-color: #FFB6C1'
                            
                            return styles
                        
                        styled_missing = missing_df.style.apply(highlight_missing, axis=1).format(format_number)
                        st.dataframe(styled_missing, use_container_width=True)
                    else:
                        st.info("No missing items found")
            
            with col2:
                with st.expander("📊 View Items with Differences", expanded=False):
                    if diff_items:
                        diff_df = pivot_table.loc[diff_items].reset_index()
                        
                        def highlight_diff(row):
                            styles = [''] * len(row)
                            
                            if 'Machinery Location' in row.index:
                                machinery_idx = row.index.get_loc('Machinery Location')
                                styles[machinery_idx] = 'background-color: #FFFF99'
                                
                                for col_name in row.index:
                                    if 'Difference' in col_name:
                                        diff_idx = row.index.get_loc(col_name)
                                        if row[col_name] != 0:
                                            styles[diff_idx] = 'background-color: #FFFF99'
                            
                            return styles
                        
                        styled_diff = diff_df.style.apply(highlight_diff, axis=1).format(format_number)
                        st.dataframe(styled_diff, use_container_width=True)
                    else:
                        st.info("No items with differences found")
            
            st.subheader("Comparison Visualization - High Difference Items")
            
            comparison_data = pivot_table[file_cols].copy()
            comparison_data['Max_Difference'] = 0
            
            for i in range(len(file_cols)):
                for j in range(i + 1, len(file_cols)):
                    diff = abs(comparison_data[file_cols[i]] - comparison_data[file_cols[j]])
                    comparison_data['Max_Difference'] = comparison_data['Max_Difference'].combine(diff, max)
            
            comparison_data = comparison_data[comparison_data['Max_Difference'] > 0]
            comparison_data = comparison_data.sort_values('Max_Difference', ascending=False).head(20)
            comparison_df = comparison_data[file_cols].reset_index()
            
            if len(comparison_df) > 0:
                fig = go.Figure()
                for file_col in file_cols:
                    fig.add_trace(go.Bar(
                        name=file_col,
                        x=comparison_df['Machinery Location'],
                        y=comparison_df[file_col]
                    ))
                
                fig.update_layout(
                    title='Top 20 Machinery Locations with Highest Differences',
                    xaxis_title='Machinery Location',
                    yaxis_title='Number of Jobs',
                    barmode='group',
                    xaxis_tickangle=-45,
                    height=600
                )
                st.plotly_chart(fig, use_container_width=True)
            else:
                st.info("No machinery with differences found")
        else:
            st.info("Upload at least 2 files to see comparison")

else:
    st.info("👆 Please upload CSV or Excel files to begin analysis")
    
    st.markdown("""
    ### Features:
    - ✅ Upload multiple CSV or Excel files
    - ✅ Automatic identification of Critical Jobs (marked with 'C')
    - ✅ Count total and critical jobs per file
    - ✅ Extract vessel name and date from data
    - ✅ Clean all columns by removing extra spaces
    - ✅ Count jobs by machinery location
    - ✅ Compare job counts across vessels
    - ✅ Show differences between vessels for each machinery type
    - ✅ Interactive visualizations and data tables
    """)
